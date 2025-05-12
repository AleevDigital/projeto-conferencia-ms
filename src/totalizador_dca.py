import pandas as pd
import numpy as np
import re
from openpyxl import load_workbook

def ler_sheet(arquivo, sheet):
    df = pd.read_excel(fr'{arquivo}', sheet_name= sheet)
    df = df.iloc[16:]
    df.columns = df.iloc[0]
    df = df.iloc[1:]
    return df

def formato_contabil(valor):
    try:
        return f"{float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except ValueError:
        return valor  # Retorna o valor original caso não seja possível formatar
    
def formato_float(valor):
    try:
        # Remove o ponto separador de milhar e substitui a vírgula por ponto
        valor_formatado = valor.replace('.', '').replace(',', '.')
        # Converte a string formatada para float
        return float(valor_formatado)
    except ValueError:
        # Retorna o valor original caso não seja possível converter
        return valor
    
def escrever_valor_abaixo(ws, wb, titulo, valor, caminho_entrada, linhas_abaixo=2):
    """
    Procura o título na coluna A e escreve o valor na coluna B, a N linhas abaixo.

    Parâmetros:
    - ws: worksheet (página da planilha aberta)
    - titulo: texto que será procurado na coluna A
    - valor: valor que será escrito
    - linhas_abaixo: quantas linhas abaixo do título o valor será escrito (default=2)
    """
    for linha in range(1, ws.max_row + 1):
        celula = ws[f"A{linha}"]
        if celula.value == titulo:
            ws[f"B{linha + linhas_abaixo}"] = valor
            break  # Para depois de encontrar
    wb.save(caminho_entrada)


def totalizador_dca(caminho_dca, planilha_conferencia):

    wb = load_workbook(planilha_conferencia)

    ws = wb['D4']
    
    #D4_00001
    #DCA ANEXO I-C
    anexo_ic = ler_sheet(caminho_dca, 'DCA-Anexo I-C')
    anexo_ic = anexo_ic.rename(columns={np.nan: 'Receitas Orçamentárias'})
    anexo_ic["Receitas Orçamentárias"] = anexo_ic["Receitas Orçamentárias"].astype(str).str.strip()
    linha_desejada = anexo_ic[anexo_ic["Receitas Orçamentárias"] == "TOTAL DAS RECEITAS (III) = (I + II)"]
    valor_anexo_ic = linha_desejada["Receitas Brutas Realizadas "].values[0]

    valor_anexo_ic = formato_contabil(valor_anexo_ic)
    valor_anexo_ic

    #DCA ANEXO I-D EMPENHADAS
    anexo_id = ler_sheet(caminho_dca, 'DCA-Anexo I-D')
    anexo_id = anexo_id.rename(columns={np.nan: 'Despesas Orçamentárias'})
    anexo_id["Despesas Orçamentárias"] = anexo_id["Despesas Orçamentárias"].astype(str).str.strip()
    linha_desejada = anexo_id[anexo_id["Despesas Orçamentárias"] == "Total Geral da Despesa"]
    valor_anexo_id_empenhada = linha_desejada["Despesas Empenhadas "].values[0]

    valor_anexo_id_empenhada = formato_contabil(valor_anexo_id_empenhada)
    valor_anexo_id_empenhada

    #DCA ANEXO I-D LIQUIDADAS
    anexo_id = ler_sheet(caminho_dca, 'DCA-Anexo I-D')
    anexo_id = anexo_id.rename(columns={np.nan: 'Despesas Orçamentárias'})
    anexo_id["Despesas Orçamentárias"] = anexo_id["Despesas Orçamentárias"].astype(str).str.strip()
    linha_desejada = anexo_id[anexo_id["Despesas Orçamentárias"] == "Total Geral da Despesa"]
    valor_anexo_id_liquidadas = linha_desejada["Despesas Liquidadas "].values[0]

    valor_anexo_id_liquidadas = formato_contabil(valor_anexo_id_liquidadas)
    valor_anexo_id_liquidadas

    #DCA ANEXO I-D PAGAS
    anexo_id = ler_sheet(caminho_dca, 'DCA-Anexo I-D')
    anexo_id = anexo_id.rename(columns={np.nan: 'Despesas Orçamentárias'})
    anexo_id["Despesas Orçamentárias"] = anexo_id["Despesas Orçamentárias"].astype(str).str.strip()
    linha_desejada = anexo_id[anexo_id["Despesas Orçamentárias"] == "Total Geral da Despesa"]
    valor_anexo_id_pagas = linha_desejada["Despesas Pagas "].values[0]

    valor_anexo_id_pagas = formato_contabil(valor_anexo_id_pagas)
    valor_anexo_id_pagas

    #DCA ANEXO I-E FUNÇÃO 01
    anexo_ie = ler_sheet(caminho_dca, 'DCA-Anexo I-E')
    anexo_ie = anexo_ie.rename(columns={np.nan: 'Despesas Por Função'})
    anexo_ie["Despesas Por Função"] = anexo_ie["Despesas Por Função"].astype(str).str.strip()
    linha_desejada = anexo_ie[anexo_ie["Despesas Por Função"] == "01 - Legislativa"]
    linha_desejada
    valor_anexo_ie_01 = linha_desejada["Despesas Empenhadas "].values[0]
    valor_anexo_ie_01 = formato_contabil(valor_anexo_ie_01) 

    #DCA ANEXO I-E FUNÇÃO 04
    anexo_ie = ler_sheet(caminho_dca, 'DCA-Anexo I-E')
    anexo_ie = anexo_ie.rename(columns={np.nan: 'Despesas Por Função'})
    anexo_ie["Despesas Por Função"] = anexo_ie["Despesas Por Função"].astype(str).str.strip()
    linha_desejada = anexo_ie[anexo_ie["Despesas Por Função"] == "04 - Administração"]
    linha_desejada
    valor_anexo_ie_04 = linha_desejada["Despesas Empenhadas "].values[0]
    valor_anexo_ie_04 = formato_contabil(valor_anexo_ie_04)


    #DCA ANEXO I-E FUNÇÃO 08
    anexo_ie = ler_sheet(caminho_dca, 'DCA-Anexo I-E')
    anexo_ie = anexo_ie.rename(columns={np.nan: 'Despesas Por Função'})
    anexo_ie["Despesas Por Função"] = anexo_ie["Despesas Por Função"].astype(str).str.strip()
    linha_desejada = anexo_ie[anexo_ie["Despesas Por Função"] == "08 - Assistência Social"]
    linha_desejada
    valor_anexo_ie_08 = linha_desejada["Despesas Empenhadas "].values[0]
    valor_anexo_ie_08 = formato_contabil(valor_anexo_ie_08)

    #DCA ANEXO I-E FUNÇÃO 09
    anexo_ie = ler_sheet(caminho_dca, 'DCA-Anexo I-E')
    anexo_ie = anexo_ie.rename(columns={np.nan: 'Despesas Por Função'})
    anexo_ie["Despesas Por Função"] = anexo_ie["Despesas Por Função"].astype(str).str.strip()
    linha_desejada = anexo_ie[anexo_ie["Despesas Por Função"] == "09 - Previdência Social"]
    linha_desejada
    valor_anexo_ie_09 = linha_desejada["Despesas Empenhadas "].values[0]
    valor_anexo_ie_09 = formato_contabil(valor_anexo_ie_09)

    #DCA ANEXO I-E FUNÇÃO 10
    anexo_ie = ler_sheet(caminho_dca, 'DCA-Anexo I-E')
    anexo_ie = anexo_ie.rename(columns={np.nan: 'Despesas Por Função'})
    anexo_ie["Despesas Por Função"] = anexo_ie["Despesas Por Função"].astype(str).str.strip()
    linha_desejada = anexo_ie[anexo_ie["Despesas Por Função"] == "10 - Saúde"]
    linha_desejada
    valor_anexo_ie_10 = linha_desejada["Despesas Empenhadas "].values[0]
    valor_anexo_ie_10 = formato_contabil(valor_anexo_ie_10)

    #DCA ANEXO I-E FUNÇÃO 12
    anexo_ie = ler_sheet(caminho_dca, 'DCA-Anexo I-E')
    anexo_ie = anexo_ie.rename(columns={np.nan: 'Despesas Por Função'})
    anexo_ie["Despesas Por Função"] = anexo_ie["Despesas Por Função"].astype(str).str.strip()
    linha_desejada = anexo_ie[anexo_ie["Despesas Por Função"] == "12 - Educação"]
    linha_desejada
    valor_anexo_ie_12 = linha_desejada["Despesas Empenhadas "].values[0]
    valor_anexo_ie_12 = formato_contabil(valor_anexo_ie_12)

    #DCA ANEXO I-E FUNÇÃO 13
    anexo_ie = ler_sheet(caminho_dca, 'DCA-Anexo I-E')
    anexo_ie = anexo_ie.rename(columns={np.nan: 'Despesas Por Função'})
    anexo_ie["Despesas Por Função"] = anexo_ie["Despesas Por Função"].astype(str).str.strip()
    linha_desejada = anexo_ie[anexo_ie["Despesas Por Função"] == "13 - Cultura"]
    linha_desejada
    valor_anexo_ie_13 = linha_desejada["Despesas Empenhadas "].values[0]
    valor_anexo_ie_13 = formato_contabil(valor_anexo_ie_13)

    #DCA ANEXO I-E FUNÇÃO 15
    anexo_ie = ler_sheet(caminho_dca, 'DCA-Anexo I-E')
    anexo_ie = anexo_ie.rename(columns={np.nan: 'Despesas Por Função'})
    anexo_ie["Despesas Por Função"] = anexo_ie["Despesas Por Função"].astype(str).str.strip()
    linha_desejada = anexo_ie[anexo_ie["Despesas Por Função"] == "15 - Urbanismo"]
    linha_desejada
    valor_anexo_ie_15 = linha_desejada["Despesas Empenhadas "].values[0]
    valor_anexo_ie_15 = formato_contabil(valor_anexo_ie_15)

    #DCA ANEXO I-E FUNÇÃO 18
    anexo_ie = ler_sheet(caminho_dca, 'DCA-Anexo I-E')
    anexo_ie = anexo_ie.rename(columns={np.nan: 'Despesas Por Função'})
    anexo_ie["Despesas Por Função"] = anexo_ie["Despesas Por Função"].astype(str).str.strip()
    linha_desejada = anexo_ie[anexo_ie["Despesas Por Função"] == "18 - Gestão Ambiental"]
    linha_desejada
    valor_anexo_ie_18 = linha_desejada["Despesas Empenhadas "].values[0]
    valor_anexo_ie_18 = formato_contabil(valor_anexo_ie_18)

    #DCA ANEXO I-E FUNÇÃO 20
    anexo_ie = ler_sheet(caminho_dca, 'DCA-Anexo I-E')
    anexo_ie = anexo_ie.rename(columns={np.nan: 'Despesas Por Função'})
    anexo_ie["Despesas Por Função"] = anexo_ie["Despesas Por Função"].astype(str).str.strip()
    linha_desejada = anexo_ie[anexo_ie["Despesas Por Função"] == "20 - Agricultura"]
    linha_desejada
    valor_anexo_ie_20 = linha_desejada["Despesas Empenhadas "].values[0]
    valor_anexo_ie_20 = formato_contabil(valor_anexo_ie_20)

    #DCA ANEXO I-E FUNÇÃO 26
    anexo_ie = ler_sheet(caminho_dca, 'DCA-Anexo I-E')
    anexo_ie = anexo_ie.rename(columns={np.nan: 'Despesas Por Função'})
    anexo_ie["Despesas Por Função"] = anexo_ie["Despesas Por Função"].astype(str).str.strip()
    linha_desejada = anexo_ie[anexo_ie["Despesas Por Função"] == "26 - Transporte"]
    linha_desejada
    valor_anexo_ie_26 = linha_desejada["Despesas Empenhadas "].values[0]
    valor_anexo_ie_26 = formato_contabil(valor_anexo_ie_26)

    #DCA ANEXO I-E FUNÇÃO 27
    anexo_ie = ler_sheet(caminho_dca, 'DCA-Anexo I-E')
    anexo_ie = anexo_ie.rename(columns={np.nan: 'Despesas Por Função'})
    anexo_ie["Despesas Por Função"] = anexo_ie["Despesas Por Função"].astype(str).str.strip()
    linha_desejada = anexo_ie[anexo_ie["Despesas Por Função"] == "27 - Desporto e Lazer"]
    linha_desejada
    valor_anexo_ie_27 = linha_desejada["Despesas Empenhadas "].values[0]
    valor_anexo_ie_27 = formato_contabil(valor_anexo_ie_27)

    #DCA ANEXO I-E FUNÇÃO 28
    anexo_ie = ler_sheet(caminho_dca, 'DCA-Anexo I-E')
    anexo_ie = anexo_ie.rename(columns={np.nan: 'Despesas Por Função'})
    anexo_ie["Despesas Por Função"] = anexo_ie["Despesas Por Função"].astype(str).str.strip()
    linha_desejada = anexo_ie[anexo_ie["Despesas Por Função"] == "28 - Encargos Especiais"]
    linha_desejada
    valor_anexo_ie_28 = linha_desejada["Despesas Empenhadas "].values[0]
    valor_anexo_ie_28 = formato_contabil(valor_anexo_ie_28)

    #DCA ANEXO I-E INTRA
    anexo_ie = ler_sheet(caminho_dca, 'DCA-Anexo I-E')
    anexo_ie = anexo_ie.rename(columns={np.nan: 'Despesas Por Função'})
    anexo_ie["Despesas Por Função"] = anexo_ie["Despesas Por Função"].astype(str).str.strip()
    linha_desejada = anexo_ie[anexo_ie["Despesas Por Função"] == "Despesas Intraorçamentárias"]
    linha_desejada
    valor_anexo_ie_intra = linha_desejada["Despesas Empenhadas "].values[0]
    valor_anexo_ie_intra = formato_contabil(valor_anexo_ie_intra)

    #DCA ANEXO I-F RPPP CANCELADOS
    anexo_if = ler_sheet(caminho_dca, 'DCA-Anexo I-F')
    anexo_if = anexo_if.rename(columns={np.nan: 'Despesas Orçamentárias'})
    anexo_if["Despesas Orçamentárias"] = anexo_if["Despesas Orçamentárias"].astype(str).str.strip()
    linha_desejada = anexo_if[anexo_if["Despesas Orçamentárias"] == "Total Despesas"]
    linha_desejada
    valor_anexo_if_cancelados = linha_desejada["Restos a Pagar Processados Cancelados "].values[0]
    valor_anexo_if_cancelados = formato_contabil(valor_anexo_if_cancelados)

    #DCA ANEXO I-F RPPP PAGOS
    anexo_if = ler_sheet(caminho_dca, 'DCA-Anexo I-F')
    anexo_if = anexo_if.rename(columns={np.nan: 'Despesas Orçamentárias'})
    anexo_if["Despesas Orçamentárias"] = anexo_if["Despesas Orçamentárias"].astype(str).str.strip()
    linha_desejada = anexo_if[anexo_if["Despesas Orçamentárias"] == "Total Despesas"]
    linha_desejada
    valor_anexo_if_pagos = linha_desejada["Restos a Pagar Processados Pagos "].values[0]
    valor_anexo_if_pagos = formato_contabil(valor_anexo_if_pagos)

    #D4_00005
    #DCA ANEXO I-F RPNPP CANCELADOS
    anexo_if = ler_sheet(caminho_dca, 'DCA-Anexo I-F')
    anexo_if = anexo_if.rename(columns={np.nan: 'Despesas Orçamentárias'})
    anexo_if["Despesas Orçamentárias"] = anexo_if["Despesas Orçamentárias"].astype(str).str.strip()
    linha_desejada = anexo_if[anexo_if["Despesas Orçamentárias"] == "Total Despesas"]
    linha_desejada
    valor_anexo_if_rnpc = linha_desejada["Restos a Pagar Não Processados Cancelados "].values[0]
    valor_anexo_if_rnpc = formato_contabil(valor_anexo_if_rnpc)

    #DCA ANEXO I-F RPNPP PAGOS
    anexo_if = ler_sheet(caminho_dca, 'DCA-Anexo I-F')
    anexo_if = anexo_if.rename(columns={np.nan: 'Despesas Orçamentárias'})
    anexo_if["Despesas Orçamentárias"] = anexo_if["Despesas Orçamentárias"].astype(str).str.strip()
    linha_desejada = anexo_if[anexo_if["Despesas Orçamentárias"] == "Total Despesas"]
    linha_desejada
    valor_anexo_if_rnp_pagos = linha_desejada["Restos a Pagar Não Processados Pagos "].values[0]
    valor_anexo_if_rnp_pagos = formato_contabil(valor_anexo_if_rnp_pagos)

    #D4_00006
    #DCA ANEXO I-G RPNPP CANCELADOS
    anexo_if = ler_sheet(caminho_dca, 'DCA-Anexo I-G')
    anexo_if = anexo_if.rename(columns={np.nan: 'Despesas Orçamentárias'})
    anexo_if["Despesas Orçamentárias"] = anexo_if["Despesas Orçamentárias"].astype(str).str.strip()
    linha_desejada = anexo_if[anexo_if["Despesas Orçamentárias"] == "Despesas Exceto Intraorçamentárias"]
    linha_desejada
    valor_anexo_ig_rnpc = linha_desejada["Restos a Pagar Não Processados Cancelados "].values[0]
    valor_anexo_ig_rnpc = formato_contabil(valor_anexo_if_rnpc)

    #DCA ANEXO I-G RPNPP PAGOS
    anexo_ig = ler_sheet(caminho_dca, 'DCA-Anexo I-G')
    anexo_ig = anexo_ig.rename(columns={np.nan: 'Despesas Por Função'})
    anexo_ig["Despesas Por Função"] = anexo_ig["Despesas Por Função"].astype(str).str.strip()
    linha_desejada = anexo_ig[anexo_ig["Despesas Por Função"] == "Despesas Exceto Intraorçamentárias"]
    linha_desejada
    valor_anexo_ig_rnp_pagos = linha_desejada["Restos a Pagar Não Processados Pagos "].values[0]
    valor_anexo_ig_rnp_pagos = formato_contabil(valor_anexo_ig_rnp_pagos)

    #D4_00007
    #DCA ANEXO I-G RPPP CANCELADOS
    anexo_ig = ler_sheet(caminho_dca, 'DCA-Anexo I-G')
    anexo_ig = anexo_ig.rename(columns={np.nan: 'Despesas Orçamentárias'})
    anexo_ig["Despesas Orçamentárias"] = anexo_ig["Despesas Orçamentárias"].astype(str).str.strip()
    linha_desejada = anexo_ig[anexo_ig["Despesas Orçamentárias"] == "Despesas Exceto Intraorçamentárias"]
    linha_desejada
    valor_anexo_ig_cancelados = linha_desejada["Restos a Pagar Processados Cancelados "].values[0]
    valor_anexo_ig_cancelados = formato_contabil(valor_anexo_ig_cancelados)

    #DCA ANEXO I-G RPPP PAGOS
    anexo_ig = ler_sheet(caminho_dca, 'DCA-Anexo I-G')
    anexo_ig = anexo_ig.rename(columns={np.nan: 'Despesas Orçamentárias'})
    anexo_ig["Despesas Orçamentárias"] = anexo_ig["Despesas Orçamentárias"].astype(str).str.strip()
    linha_desejada = anexo_ig[anexo_ig["Despesas Orçamentárias"] == "Despesas Exceto Intraorçamentárias"]
    linha_desejada
    valor_anexo_ig_pagos = linha_desejada["Restos a Pagar Processados Pagos "].values[0]
    valor_anexo_ig_pagos = formato_contabil(valor_anexo_ig_pagos)

    #D4_00008
    #DCA ANEXO I-C  ALIENAÇÃO DE ATIVOS
    receitas_estaduais = ler_sheet(caminho_dca, 'DCA-Anexo I-C')
    receitas_estaduais = receitas_estaduais.rename(columns={np.nan: 'Receitas Orçamentárias'})
    receitas_estaduais["Receitas Orçamentárias"] = receitas_estaduais["Receitas Orçamentárias"].astype(str).str.strip()
    linha_desejada = receitas_estaduais[receitas_estaduais["Receitas Orçamentárias"] == "2.2.0.0.00.0.0 - Alienação de Bens"]
    linha_desejada
    valor_receitas_estaduais = linha_desejada["Receitas Brutas Realizadas "].values[0]
    valor_receitas_estaduais = formato_contabil(valor_receitas_estaduais)


    #D4_00009
    #DCA ANEXO I-C  RECEITAS COM TRIBUTOS ESTADUAIS
    receitas_estaduais = ler_sheet(caminho_dca, 'DCA-Anexo I-C')
    receitas_estaduais = receitas_estaduais.rename(columns={np.nan: 'Receitas Orçamentárias'})
    receitas_estaduais["Receitas Orçamentárias"] = receitas_estaduais["Receitas Orçamentárias"].astype(str).str.strip()
    linha_desejada = receitas_estaduais[receitas_estaduais["Receitas Orçamentárias"] == "1.1.0.0.00.0.0 - Impostos, Taxas e Contribuições de Melhoria"]
    linha_desejada
    valor_receitas_estaduais = linha_desejada["Receitas Brutas Realizadas "].values[0]
    valor_receitas_estaduais = formato_contabil(valor_receitas_estaduais)

    # #D4_00010
    #DCA ANEXO I-C  RECEITAS COM TRIBUTOS MUNICIPAIS
    receitas_estaduais = ler_sheet(caminho_dca, 'DCA-Anexo I-C')
    receitas_estaduais = receitas_estaduais.rename(columns={np.nan: 'Receitas Orçamentárias'})
    receitas_estaduais["Receitas Orçamentárias"] = receitas_estaduais["Receitas Orçamentárias"].astype(str).str.strip()
    linha_desejada = receitas_estaduais[receitas_estaduais["Receitas Orçamentárias"] == "1.1.0.0.00.0.0 - Impostos, Taxas e Contribuições de Melhoria"]
    linha_desejada
    valor_receitas_estaduais = linha_desejada["Receitas Brutas Realizadas "].values[0]
    valor_receitas_estaduais = formato_contabil(valor_receitas_estaduais)

    #D4_00012
    #DCA ANEXO I-C  TRANSFERENCIAS MUNICIPAIS
    transferencias_municipais = ler_sheet(caminho_dca, 'DCA-Anexo I-C')
    transferencias_municipais = transferencias_municipais.rename(columns={np.nan: 'Receitas Orçamentárias'})
    transferencias_municipais["Receitas Orçamentárias"] = transferencias_municipais["Receitas Orçamentárias"].astype(str).str.strip()
    linha_desejada = transferencias_municipais[transferencias_municipais["Receitas Orçamentárias"] == "1.7.1.0.00.0.0 - Transferências da União e de suas Entidades"]
    linha_desejada
    valor_transferencias_municipais = linha_desejada["Receitas Brutas Realizadas "].values[0]
    valor_transferencias_municipais = formato_contabil(valor_transferencias_municipais)

    #D4_00013
    #DCA ANEXO I-C  RECEITAS COM TRIBUTOS ESTADUAIS ICMS
    cota_icms = ler_sheet(caminho_dca, 'DCA-Anexo I-C')
    cota_icms = cota_icms.rename(columns={np.nan: 'Receitas Orçamentárias'})
    cota_icms["Receitas Orçamentárias"] = cota_icms["Receitas Orçamentárias"].astype(str).str.strip()
    linha_desejada = cota_icms[cota_icms["Receitas Orçamentárias"] == "1.7.2.1.50.0.0 - Cota-Parte do ICMS"]
    linha_desejada
    valor_cota_icms = linha_desejada["Receitas Brutas Realizadas "].values[0]
    valor_cota_icms = formato_contabil(valor_cota_icms)

    #D4_00013
    #DCA ANEXO I-C  RECEITAS COM TRIBUTOS ESTADUAIS IPVA
    cota_ipva = ler_sheet(caminho_dca, 'DCA-Anexo I-C')
    cota_ipva = cota_ipva.rename(columns={np.nan: 'Receitas Orçamentárias'})
    cota_ipva["Receitas Orçamentárias"] = cota_ipva["Receitas Orçamentárias"].astype(str).str.strip()
    linha_desejada = cota_ipva[cota_ipva["Receitas Orçamentárias"] == "1.7.2.1.51.0.0 - Cota-Parte do IPVA"]
    linha_desejada
    valor_cota_ipva = linha_desejada["Receitas Brutas Realizadas "].values[0]
    valor_cota_ipva = formato_contabil(valor_cota_ipva)

    #D4_00014
    #DCA ANEXO I-C  RECEITAS COM TRIBUTOS ESTADUAIS ICMS E IPVA
    cota_iptu = ler_sheet(caminho_dca, 'DCA-Anexo I-C')
    cota_iptu = cota_iptu.rename(columns={np.nan: 'Receitas Orçamentárias'})
    cota_iptu["Receitas Orçamentárias"] = cota_iptu["Receitas Orçamentárias"].astype(str).str.strip()
    linha_desejada = cota_iptu[cota_iptu["Receitas Orçamentárias"] == "1.1.1.0.00.0.0 - Impostos"]
    linha_desejada
    valor_cota_iptu = linha_desejada["Receitas Brutas Realizadas "].values[0]
    valor_cota_iptu = formato_contabil(valor_cota_iptu)

    #D4_00019
    #DCA ANEXO I-C  RECEITAS COM TRIBUTOS ESTADUAIS ICMS E IPVA
    despesas_capital = ler_sheet(caminho_dca, 'DCA-Anexo I-D')
    despesas_capital = despesas_capital.rename(columns={np.nan: 'Receitas Orçamentárias'})
    despesas_capital["Receitas Orçamentárias"] = despesas_capital["Receitas Orçamentárias"].astype(str).str.strip()
    linha_desejada = despesas_capital[despesas_capital["Receitas Orçamentárias"] == "4.0.00.00.00 - Despesas de Capital"]
    linha_desejada
    valor_despesas_capital = linha_desejada["Despesas Empenhadas "].values[0]
    valor_despesas_capital = formato_contabil(valor_despesas_capital)


    #D4_00016
    #DCA ANEXO I-C  TRANSFERÊNCIAS CONSTITUCIONAIS MUNICIPAIS
    codigos_desejados = [
        "1.7.1.1.51",
        "1.7.1.1.52",
        "1.7.2.1.50",
        "1.7.2.1.51",
        "1.7.1.5.00",
        "1.7.5.1.00"
    ]

    constitucionais_municipais = ler_sheet(caminho_dca, 'DCA-Anexo I-C')
    constitucionais_municipais = constitucionais_municipais.rename(columns={np.nan: 'Receitas Orçamentárias'})
    constitucionais_municipais["Receitas Orçamentárias"] = constitucionais_municipais["Receitas Orçamentárias"].astype(str).str.strip()
    padrao_regex = "|".join(map(re.escape, codigos_desejados))
    linhas_filtradas = constitucionais_municipais[constitucionais_municipais["Receitas Orçamentárias"].str.contains(padrao_regex, na=False)]
    valor_constitucionais = linhas_filtradas["Receitas Brutas Realizadas "].sum()
    valor_constitucionais = formato_contabil(valor_constitucionais)

    escrever_valor_abaixo(ws,wb,"D4_00001",valor_anexo_ic,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00002",valor_anexo_id_empenhada,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00002",valor_anexo_id_liquidadas,planilha_conferencia,3)
    escrever_valor_abaixo(ws,wb,"D4_00002",valor_anexo_id_pagas,planilha_conferencia,4)
    escrever_valor_abaixo(ws,wb,"D4_00003",valor_anexo_ie_01,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00003",valor_anexo_ie_04,planilha_conferencia,3)
    escrever_valor_abaixo(ws,wb,"D4_00003",valor_anexo_ie_08,planilha_conferencia,4)
    escrever_valor_abaixo(ws,wb,"D4_00003",valor_anexo_ie_09,planilha_conferencia,5)
    escrever_valor_abaixo(ws,wb,"D4_00003",valor_anexo_ie_10,planilha_conferencia,6)
    escrever_valor_abaixo(ws,wb,"D4_00003",valor_anexo_ie_12,planilha_conferencia,7)
    escrever_valor_abaixo(ws,wb,"D4_00003",valor_anexo_ie_13,planilha_conferencia,8)
    escrever_valor_abaixo(ws,wb,"D4_00003",valor_anexo_ie_15,planilha_conferencia,9)
    escrever_valor_abaixo(ws,wb,"D4_00003",valor_anexo_ie_18,planilha_conferencia,10)
    escrever_valor_abaixo(ws,wb,"D4_00003",valor_anexo_ie_20,planilha_conferencia,11)
    escrever_valor_abaixo(ws,wb,"D4_00003",valor_anexo_ie_26,planilha_conferencia,12)
    escrever_valor_abaixo(ws,wb,"D4_00003",valor_anexo_ie_27,planilha_conferencia,13)
    escrever_valor_abaixo(ws,wb,"D4_00003",valor_anexo_ie_28,planilha_conferencia,14)
    escrever_valor_abaixo(ws,wb,"D4_00004",valor_anexo_ie_intra,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00005",valor_anexo_if_cancelados,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00005",valor_anexo_if_pagos,planilha_conferencia,3)
    escrever_valor_abaixo(ws,wb,"D4_00005",valor_anexo_if_rnpc,planilha_conferencia,4)
    escrever_valor_abaixo(ws,wb,"D4_00005",valor_anexo_if_rnp_pagos,planilha_conferencia,5)
    escrever_valor_abaixo(ws,wb,"D4_00006",valor_anexo_ig_rnpc,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00006",valor_anexo_ig_rnp_pagos,planilha_conferencia,3)
    escrever_valor_abaixo(ws,wb,"D4_00007",valor_anexo_ig_cancelados,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00007",valor_anexo_ig_pagos,planilha_conferencia,3)
    escrever_valor_abaixo(ws,wb,"D4_00008",valor_receitas_estaduais,planilha_conferencia)
    #escrever_valor_abaixo(ws,wb,"D4_00009",valor_receitas_estaduais,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00010",valor_receitas_estaduais,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00012",valor_transferencias_municipais,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00014",valor_cota_iptu,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00019",valor_despesas_capital,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00016",valor_constitucionais,planilha_conferencia)





    











