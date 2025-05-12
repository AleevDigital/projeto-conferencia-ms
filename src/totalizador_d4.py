import pandas as pd
import os
from openpyxl import load_workbook

def escrever_valor_abaixo(ws, wb, titulo, valor, caminho_entrada, linhas_abaixo=2):
    """
    Procura o título na coluna A e escreve o valor na coluna B, a N linhas abaixo.
    
    Parâmetros:
    - ws: worksheet (página da planilha aberta)
    - titulo: texto que será procurado na coluna A
    - valor: valor que será escrito
    - linhas_abaixo: quantas linhas abaixo do título o valor será escrito (default=2)
    """
    for linha in range(1, ws.max_row + 1):
        celula = ws[f"A{linha}"]
        if celula.value == titulo:
            ws[f"B{linha + linhas_abaixo}"] = valor
            break  # Para depois de encontrar
    wb.save(caminho_entrada)


def formato_contabil(valor):
    try:
        return f"{float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except ValueError:
        return valor  # Retorna o valor original caso não seja possível formatar
    
def formato_float(valor):
    try:
        # Remove o ponto separador de milhar e substitui a vírgula por ponto
        valor_formatado = valor.replace('.', '').replace(',', '.')
        # Converte a string formatada para float
        return float(valor_formatado)
    except ValueError:
        # Retorna o valor original caso não seja possível converter
        return valor
    
def totalizador_d4(arquivo, planilha_conferencia):

    nome_arquivo = os.path.basename(planilha_conferencia)
    wb = load_workbook(planilha_conferencia)

    ws = wb['D4']

    #D4_00020
    arquivo['Conta Contábil'] = arquivo['Conta Contábil'].astype(str)
    filtro = (arquivo['Conta Contábil'].str.startswith('6212') | arquivo['Conta Contábil'].str.startswith('6213'))
    conta_6212_6213 = arquivo[filtro]
    conta_6212_6213_total= conta_6212_6213['Saldo Final'].sum()
    conta_6212_6213_total = formato_contabil(conta_6212_6213_total)

    #D4_00022
    contas_desejadas = ["621200000", "621300000"]
    conta_6212_6213 = arquivo[arquivo['Conta Contábil'].isin(contas_desejadas)]
    contas_111_4 = conta_6212_6213[conta_6212_6213['Informações Complementares 4'].astype(str).str.startswith('111')]
    tipos_para_ignorar = ["Complemento da Fonte de Recursos ou Destinação de Recursos"]
    contas_111_3 = conta_6212_6213[
        conta_6212_6213['Informações Complementares 3'].astype(str).str.startswith('111') &
        ~conta_6212_6213['Tipo de Informação 3'].isin(tipos_para_ignorar)]
    contas_111 = pd.concat([contas_111_4, contas_111_3], axis=0, ignore_index=True)
    msc_dezembro_111 = contas_111['Saldo Final'].sum()
    msc_dezembro_111 = formato_contabil(msc_dezembro_111)

    #D4_00024
    contas_desejadas = ["621200000", "621300000"]
    conta_6212_6213 = arquivo[arquivo['Conta Contábil'].isin(contas_desejadas)]
    padroes_procurados = ("171151", "171152", "172150", "172151", "1715", "1751")
    conta_info4 = conta_6212_6213[
        conta_6212_6213['Informações Complementares 4'].astype(str).str.startswith(padroes_procurados)]
    conta_info3 = conta_6212_6213[
        conta_6212_6213['Informações Complementares 3'].astype(str).str.startswith(padroes_procurados)]
    conta_info_3_4 = pd.concat([conta_info4, conta_info3], axis=0, ignore_index=True)
    saldo_total = conta_info_3_4['Saldo Final'].sum()
    saldo_total_formatado = formato_contabil(saldo_total)

    #DESPESAS MSC DEZEMBRO D4_00025 LINHA 1
    contas_desejadas = arquivo['Conta Contábil'].astype(str).str.startswith('62213')
    conta_62213 = arquivo[contas_desejadas]
    conta_62213_total_25 = conta_62213['Saldo Final'].sum()
    conta_62213_total_25 = formato_contabil(conta_62213_total_25)

    # EXCETO 01 02 E 05 D4_00025 LINHA 2
    contas_desejadas = arquivo[arquivo['Conta Contábil'].astype(str).str.startswith('62213')]
    condicao = contas_desejadas['Conta Contábil'].isin(["622130100","622130200","622130500"])
    contas_62213_especificas = contas_desejadas[~condicao]
    contas_62213_especificas_25 = contas_62213_especificas['Saldo Final'].sum()
    contas_62213_especificas_25 = formato_contabil(contas_62213_especificas_25)

    #622130400 D4_00025 LINHA 3
    conta_desejada = "622130400"
    conta_6221304 = arquivo[arquivo['Conta Contábil'] == conta_desejada]
    conta_6221304 = conta_6221304['Saldo Final'].sum()
    conta_6221304 = formato_contabil(conta_6221304)

    # 6221305 D4_00026
    contas_desejadas = arquivo[arquivo['Conta Contábil'] == "622130500"]
    conta_6221305 = contas_desejadas['Saldo Final'].sum()
    conta_6221305 = formato_contabil(conta_6221305)
    conta_6221305

    #D4_00029 L1
    filtro_62213_l1 = arquivo['Conta Contábil'].astype(str).str.startswith('62213')
    dados_62213_l1 = arquivo[filtro_62213_l1]
    filtro_info2_10_l1 = dados_62213_l1['Informações Complementares 2'].astype(str).str.match(r'^(9|09)')
    dados_62213_l1 = dados_62213_l1[filtro_info2_10_l1]
    info5_l1 = dados_62213_l1['Informações Complementares 5'].astype(str)
    condicao_retirar_l1 = (info5_l1.str.len() >= 4) & (info5_l1.str[2] == '9') & (info5_l1.str[3] == '1')
    dados_62213_filtrados_l1 = dados_62213_l1[~condicao_retirar_l1]
    saldo_l1 = dados_62213_filtrados_l1['Saldo Final'].sum()
    saldo_l19_formatado = formato_contabil(saldo_l1)

    #D4_00029 L2
    filtro_62213_l2 = arquivo['Conta Contábil'].astype(str).str.startswith('62213')
    dados_62213_l2 = arquivo[filtro_62213_l2]
    filtro_info2_10_l2 = dados_62213_l2['Informações Complementares 2'].astype(str).str.match(r'^(9|09)')
    dados_62213_l2 = dados_62213_l2[filtro_info2_10_l2]
    contas_para_remover_l2 = ["622130100", "622130200", "622130500"]
    condicao_remover_contas_l2 = dados_62213_l2['Conta Contábil'].isin(contas_para_remover_l2)
    dados_62213_sem_especificas_l2 = dados_62213_l2[~condicao_remover_contas_l2]
    info5_l2 = dados_62213_sem_especificas_l2['Informações Complementares 5'].astype(str)
    condicao_retirar_l2 = (info5_l2.str.len() >= 4) & (info5_l2.str[2] == '9') & (info5_l2.str[3] == '1')
    dados_62213_filtrados_l2 = dados_62213_sem_especificas_l2[~condicao_retirar_l2]
    saldo_l2 = dados_62213_filtrados_l2['Saldo Final'].sum()
    saldo_l29_formatado = formato_contabil(saldo_l2)

    #D4_00030 L1

    filtro_62213_l1 = arquivo['Conta Contábil'].astype(str).str.startswith('62213')
    dados_62213_l1 = arquivo[filtro_62213_l1]
    filtro_info2_10_l1 = dados_62213_l1['Informações Complementares 2'].astype(str).str.startswith('10')
    dados_62213_l1 = dados_62213_l1[filtro_info2_10_l1]
    condicao_1031 = dados_62213_l1['Informações Complementares 2'] == "1031"
    dados_62213_l1 = dados_62213_l1[~condicao_1031]
    info5_l1 = dados_62213_l1['Informações Complementares 5'].astype(str)
    condicao_retirar_l1 = (info5_l1.str.len() >= 4) & (info5_l1.str[2] == '9') & (info5_l1.str[3] == '1')
    dados_62213_filtrados_l1 = dados_62213_l1[~condicao_retirar_l1]
    saldo_l1 = dados_62213_filtrados_l1['Saldo Final'].sum()
    saldo_l110_formatado = formato_contabil(saldo_l1)

    #D4_00030 L2
    filtro_62213_l2 = arquivo['Conta Contábil'].astype(str).str.startswith('62213')
    dados_62213_l2 = arquivo[filtro_62213_l2]
    filtro_info2_10_l2 = dados_62213_l2['Informações Complementares 2'].astype(str).str.startswith('10')
    dados_62213_l2 = dados_62213_l2[filtro_info2_10_l2]
    condicao_1031 = dados_62213_l2['Informações Complementares 2'] == "1031"
    dados_62213_l2 = dados_62213_l2[~condicao_1031]
    contas_para_remover_l2 = ["622130100", "622130200", "622130500"]
    condicao_remover_contas_l2 = dados_62213_l2['Conta Contábil'].isin(contas_para_remover_l2)
    dados_62213_sem_especificas_l2 = dados_62213_l2[~condicao_remover_contas_l2]
    info5_l2 = dados_62213_sem_especificas_l2['Informações Complementares 5'].astype(str)
    condicao_retirar_l2 = (info5_l2.str.len() >= 4) & (info5_l2.str[2] == '9') & (info5_l2.str[3] == '1')
    dados_62213_filtrados_l2 = dados_62213_sem_especificas_l2[~condicao_retirar_l2]
    saldo_l2 = dados_62213_filtrados_l2['Saldo Final'].sum()
    saldo_l210_formatado = formato_contabil(saldo_l2)

    #D4_00031 L1
    filtro_62213_l1 = arquivo['Conta Contábil'].astype(str).str.startswith('62213')
    dados_62213_l1 = arquivo[filtro_62213_l1]
    filtro_info2_10_l1 = dados_62213_l1['Informações Complementares 2'].astype(str).str.startswith('12')
    dados_62213_l1 = dados_62213_l1[filtro_info2_10_l1]
    info5_l1 = dados_62213_l1['Informações Complementares 5'].astype(str)
    condicao_retirar_l1 = (info5_l1.str.len() >= 4) & (info5_l1.str[2] == '9') & (info5_l1.str[3] == '1')
    dados_62213_filtrados_l1 = dados_62213_l1[~condicao_retirar_l1]
    saldo_l1 = dados_62213_filtrados_l1['Saldo Final'].sum()
    saldo_l112_formatado = formato_contabil(saldo_l1)

    #D4_00031 L2
    filtro_62213_l2 = arquivo['Conta Contábil'].astype(str).str.startswith('62213')
    dados_62213_l2 = arquivo[filtro_62213_l2]
    filtro_info2_10_l2 = dados_62213_l2['Informações Complementares 2'].astype(str).str.startswith('12')
    dados_62213_l2 = dados_62213_l2[filtro_info2_10_l2]
    contas_para_remover_l2 = ["622130100", "622130200", "622130500"]
    condicao_remover_contas_l2 = dados_62213_l2['Conta Contábil'].isin(contas_para_remover_l2)
    dados_62213_sem_especificas_l2 = dados_62213_l2[~condicao_remover_contas_l2]
    info5_l2 = dados_62213_sem_especificas_l2['Informações Complementares 5'].astype(str)
    condicao_retirar_l2 = (info5_l2.str.len() >= 4) & (info5_l2.str[2] == '9') & (info5_l2.str[3] == '1')
    dados_62213_filtrados_l2 = dados_62213_sem_especificas_l2[~condicao_retirar_l2]
    saldo_l2 = dados_62213_filtrados_l2['Saldo Final'].sum()
    saldo_l212_formatado = formato_contabil(saldo_l2)

    #D4_00032
    conta_62213_total_float = formato_float(conta_62213_total_25)
    saldo_l19_formatado_float = formato_float(saldo_l19_formatado)
    saldo_l110_formatado_float = formato_float(saldo_l110_formatado)
    saldo_l112_formatado_float = formato_float(saldo_l112_formatado)

    contas_62213_especificas_float = formato_float(contas_62213_especificas_25)
    saldo_l29_formatado_float = formato_float(saldo_l29_formatado)
    saldo_l210_formatado_float = formato_float(saldo_l210_formatado)
    saldo_l212_formatado_float = formato_float(saldo_l212_formatado)

    conta_72_l1  = conta_62213_total_float-saldo_l19_formatado_float-saldo_l110_formatado_float-saldo_l112_formatado_float
    conta_72_l1 = formato_contabil(conta_72_l1)

    conta_72_l2  = contas_62213_especificas_float-saldo_l29_formatado_float-saldo_l210_formatado_float-saldo_l212_formatado_float
    conta_72_l2 = formato_contabil(conta_72_l2)

    #D4_00033
    contas_desejadas = arquivo['Conta Contábil'].astype(str).str.startswith('62213')
    conta_62213 = arquivo[contas_desejadas]
    info5_l1 = conta_62213['Informações Complementares 5'].astype(str)
    condicao_colocar_l1 = (info5_l1.str.len() >= 4) & (info5_l1.str[2] == '9') & (info5_l1.str[3] == '1')
    conta_62213 = conta_62213[condicao_colocar_l1]
    conta_62213_total_73 = conta_62213['Saldo Inicial'].sum()
    conta_62213_total_73 = formato_contabil(conta_62213_total_73)

    # EXCETO 01 02 E 05 D2_00049 LINHA 2
    contas_desejadas = arquivo[arquivo['Conta Contábil'].astype(str).str.startswith('62213')]
    condicao = contas_desejadas['Conta Contábil'].isin(["622130100", "622130200", "622130500"])
    contas_62213_especificas = contas_desejadas[~condicao]
    info5_l2 = contas_62213_especificas['Informações Complementares 5'].astype(str)
    condicao_colocar_l2 = (info5_l2.str.len() >= 4) & (info5_l2.str[2] == '9') & (info5_l2.str[3] == '1')
    contas_62213_especificas = contas_62213_especificas[condicao_colocar_l2]
    contas_62213_especificas_73 = contas_62213_especificas['Saldo Inicial'].sum()
    contas_62213_especificas_73 = formato_contabil(contas_62213_especificas_73)


    #D4_00034 LINHA 1
    contas_desejadas = arquivo[arquivo['Conta Contábil'].astype(str).str.startswith('6322')]
    conta_6322  = contas_desejadas['Saldo Final'].sum()
    conta_6322 = formato_contabil(conta_6322)
    conta_6322

    #D4_00034 LINHA 2
    contas_desejadas = arquivo[arquivo['Conta Contábil'].astype(str).str.startswith('6314')]
    conta_6314  = contas_desejadas['Saldo Final'].sum()
    conta_6314 = formato_contabil(conta_6314)
    conta_6314

    #D4_00035 e #D400036
    contas_desejadas = arquivo[arquivo['Conta Contábil'].astype(str).str.startswith('111')]
    conta_contabil_111 = contas_desejadas['Saldo Final'].sum()
    conta_contabil_111 = formato_contabil(conta_contabil_111)
    conta_contabil_111

    escrever_valor_abaixo(ws,wb,"D4_00020",conta_6212_6213_total,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00022",msc_dezembro_111,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00024",saldo_total_formatado,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00025",conta_62213_total_25,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00025",contas_62213_especificas_25,planilha_conferencia,3)
    escrever_valor_abaixo(ws,wb,"D4_00025",conta_6221304,planilha_conferencia,4)
    escrever_valor_abaixo(ws,wb,"D4_00026",conta_6221305,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00029",saldo_l19_formatado,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00029",saldo_l29_formatado,planilha_conferencia,3)
    escrever_valor_abaixo(ws,wb,"D4_00030",saldo_l110_formatado,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00030",saldo_l210_formatado,planilha_conferencia,3)
    escrever_valor_abaixo(ws,wb,"D4_00031",saldo_l112_formatado,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00031",saldo_l212_formatado,planilha_conferencia,3)
    escrever_valor_abaixo(ws,wb,"D4_00032",conta_72_l1,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00032",conta_72_l2,planilha_conferencia,3)
    escrever_valor_abaixo(ws,wb,"D4_00033",conta_62213_total_73,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00033",contas_62213_especificas_73,planilha_conferencia,3)
    escrever_valor_abaixo(ws,wb,"D4_00034",conta_6322,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00034",conta_6314,planilha_conferencia,3)
    escrever_valor_abaixo(ws,wb,"D4_00035",conta_contabil_111,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D4_00036",conta_contabil_111,planilha_conferencia)