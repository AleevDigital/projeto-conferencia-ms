import pandas as pd
import os
from openpyxl import load_workbook


def escrever_valor_abaixo(ws, wb, titulo, valor, caminho_entrada, linhas_abaixo=2):
    """
    Procura o título na coluna A e escreve o valor na coluna B, a N linhas abaixo.
    
    Parâmetros:
    - ws: worksheet (página da planilha aberta)
    - titulo: texto que será procurado na coluna A
    - valor: valor que será escrito
    - linhas_abaixo: quantas linhas abaixo do título o valor será escrito (default=2)
    """
    for linha in range(1, ws.max_row + 1):
        celula = ws[f"A{linha}"]
        if celula.value == titulo:
            ws[f"B{linha + linhas_abaixo}"] = valor
            break  # Para depois de encontrar
    wb.save(caminho_entrada)


def formato_contabil(valor):
    try:
        return f"{float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except ValueError:
        return valor  # Retorna o valor original caso não seja possível formatar
    
def formato_float(valor):
    try:
        # Remove o ponto separador de milhar e substitui a vírgula por ponto
        valor_formatado = valor.replace('.', '').replace(',', '.')
        # Converte a string formatada para float
        return float(valor_formatado)
    except ValueError:
        # Retorna o valor original caso não seja possível converter
        return valor

def totalizador_d2(arquivo, planilha_conferencia):
    nome_arquivo = os.path.basename(planilha_conferencia)
    wb = load_workbook(planilha_conferencia)

    ws = wb['D2']

    # D2_00044
    arquivo['Conta Contábil'] = arquivo['Conta Contábil'].astype(str)
    filtro = (
        arquivo['Conta Contábil'].str.startswith('6212') |
        arquivo['Conta Contábil'].str.startswith('6213')
    )
    conta_6212_6213 = arquivo[filtro]
    conta_6212_6213_total= conta_6212_6213['Saldo Inicial'].sum()
    conta_6212_6213_total = formato_contabil(conta_6212_6213_total)

    #D2_00045
    contas_desejadas = ["621200000", "621300000"]
    conta_6212_6213 = arquivo[arquivo['Conta Contábil'].isin(contas_desejadas)]
    contas_172_4 = conta_6212_6213[
        conta_6212_6213['Informações Complementares 4'].astype(str).str.startswith('172')]
    contas_172_3 = conta_6212_6213[
        conta_6212_6213['Informações Complementares 3'].astype(str).str.startswith('172')]
    
    contas_172 = pd.concat([contas_172_4, contas_172_3], axis=0, ignore_index=True)
    msc_dezembro_172 = contas_172['Saldo Inicial'].sum()
    msc_dezembro_172 = formato_contabil(msc_dezembro_172)

    # D2_00046
    contas_desejadas = ["621200000", "621300000"]
    conta_6212_6213 = arquivo[arquivo['Conta Contábil'].isin(contas_desejadas)]
    contas_111_4 = conta_6212_6213[
        conta_6212_6213['Informações Complementares 4'].astype(str).str.startswith('111')]
    tipos_para_ignorar = ["Complemento da Fonte de Recursos ou Destinação de Recursos"]

    contas_111_3 = conta_6212_6213[
        conta_6212_6213['Informações Complementares 3'].astype(str).str.startswith('111') &
        ~conta_6212_6213['Tipo de Informação 3'].isin(tipos_para_ignorar)]
    contas_111 = pd.concat([contas_111_4, contas_111_3], axis=0, ignore_index=True)
    msc_dezembro_111 = contas_111['Saldo Inicial'].sum()
    msc_dezembro_111 = formato_contabil(msc_dezembro_111)

    #D2_00047
    contas_desejadas = ["621200000", "621300000"]
    conta_6212_6213 = arquivo[arquivo['Conta Contábil'].isin(contas_desejadas)]
    padroes_procurados = ("1712", "175")
    conta_info4 = conta_6212_6213[
        conta_6212_6213['Informações Complementares 4'].astype(str).str.startswith(padroes_procurados)]
    conta_info3 = conta_6212_6213[
        conta_6212_6213['Informações Complementares 3'].astype(str).str.startswith(padroes_procurados)]
    conta_info_3_4 = pd.concat([conta_info4, conta_info3], axis=0, ignore_index=True)
    saldo_total = conta_info_3_4['Saldo Inicial'].sum()
    saldo_total_formatado = formato_contabil(saldo_total)

    #D2_00048
    contas_desejadas = ["621200000", "621300000"]
    conta_6212_6213 = arquivo[arquivo['Conta Contábil'].isin(contas_desejadas)]
    padroes_procurados = ("171151", "171152", "172150", "172151", "1715", "1751")
    conta_info4 = conta_6212_6213[
        conta_6212_6213['Informações Complementares 4'].astype(str).str.startswith(padroes_procurados)]
    conta_info3 = conta_6212_6213[
        conta_6212_6213['Informações Complementares 3'].astype(str).str.startswith(padroes_procurados)]
    conta_info_3_4 = pd.concat([conta_info4, conta_info3], axis=0, ignore_index=True)
    saldo_total = conta_info_3_4['Saldo Inicial'].sum()
    saldo_total_formatado_48 = formato_contabil(saldo_total)

    #DESPESAS MSC DEZEMBRO D2_00049 LINHA 1
    contas_desejadas = arquivo['Conta Contábil'].astype(str).str.startswith('62213')
    conta_62213 = arquivo[contas_desejadas]
    conta_62213_total_49 = conta_62213['Saldo Inicial'].sum()
    conta_62213_total_49 = formato_contabil(conta_62213_total_49)

    # EXCETO 01 02 E 05 D2_00049 LINHA 2
    contas_desejadas = arquivo[arquivo['Conta Contábil'].astype(str).str.startswith('62213')]
    condicao = contas_desejadas['Conta Contábil'].isin(["622130100","622130200","622130500"])
    contas_62213_especificas = contas_desejadas[~condicao]
    contas_62213_especificas_49 = contas_62213_especificas['Saldo Inicial'].sum()
    contas_62213_especificas_49 = formato_contabil(contas_62213_especificas_49)

    #622130400 D2_00049 LINHA 3
    conta_desejada = "622130400"
    conta_6221304 = arquivo[arquivo['Conta Contábil'] == conta_desejada]
    conta_6221304_49 = conta_6221304['Saldo Inicial'].sum()
    conta_6221304_49 = formato_contabil(conta_6221304_49)

    # 6221307 D2_00050 L1
    contas_desejadas = arquivo[arquivo['Conta Contábil'] == "622130700"]
    conta_6221307 = contas_desejadas['Saldo Inicial'].sum()
    conta_6221307 = formato_contabil(conta_6221307)

    # 6221305 D2_00050 L2
    contas_desejadas = arquivo[arquivo['Conta Contábil'] == "622130500"]
    conta_6221305 = contas_desejadas['Saldo Inicial'].sum()
    conta_6221305 = formato_contabil(conta_6221305)

    #D2_00058 L1
    contas_desejadas = arquivo['Conta Contábil'].astype(str).str.startswith('4522')
    conta_4522 = arquivo[contas_desejadas]
    conta_4522_total = conta_4522['Saldo Inicial'].sum()
    conta_4522_total = formato_contabil(conta_4522_total)
    
    #D2_00058 L2
    contas_desejadas = arquivo['Conta Contábil'].astype(str).str.startswith('3522')
    conta_3522 = arquivo[contas_desejadas]
    conta_3522_total = conta_3522['Saldo Inicial'].sum()
    conta_3522_total = formato_contabil(conta_3522_total)

    #D2_00069 L1
    filtro_62213_l1 = arquivo['Conta Contábil'].astype(str).str.startswith('62213')
    dados_62213_l1 = arquivo[filtro_62213_l1]
    filtro_info2_10_l1 = dados_62213_l1['Informações Complementares 2'].astype(str).str.match(r'^(9|09)')
    dados_62213_l1 = dados_62213_l1[filtro_info2_10_l1]
    info5_l1 = dados_62213_l1['Informações Complementares 5'].astype(str)
    condicao_retirar_l1 = (info5_l1.str.len() >= 4) & (info5_l1.str[2] == '9') & (info5_l1.str[3] == '1')
    dados_62213_filtrados_l1 = dados_62213_l1[~condicao_retirar_l1]
    saldo_l1 = dados_62213_filtrados_l1['Saldo Inicial'].sum()
    saldo_l19_formatado = formato_contabil(saldo_l1)

    #D2_00069 L2
    filtro_62213_l2 = arquivo['Conta Contábil'].astype(str).str.startswith('62213')
    dados_62213_l2 = arquivo[filtro_62213_l2]
    filtro_info2_10_l2 = dados_62213_l2['Informações Complementares 2'].astype(str).str.match(r'^(9|09)')
    dados_62213_l2 = dados_62213_l2[filtro_info2_10_l2]
    contas_para_remover_l2 = ["622130100", "622130200", "622130500"]
    condicao_remover_contas_l2 = dados_62213_l2['Conta Contábil'].isin(contas_para_remover_l2)
    dados_62213_sem_especificas_l2 = dados_62213_l2[~condicao_remover_contas_l2]
    info5_l2 = dados_62213_sem_especificas_l2['Informações Complementares 5'].astype(str)
    condicao_retirar_l2 = (info5_l2.str.len() >= 4) & (info5_l2.str[2] == '9') & (info5_l2.str[3] == '1')
    dados_62213_filtrados_l2 = dados_62213_sem_especificas_l2[~condicao_retirar_l2]
    saldo_l2 = dados_62213_filtrados_l2['Saldo Inicial'].sum()
    saldo_l29_formatado = formato_contabil(saldo_l2)

    #D2_00070 L1
    filtro_62213_l1 = arquivo['Conta Contábil'].astype(str).str.startswith('62213')
    dados_62213_l1 = arquivo[filtro_62213_l1]
    filtro_info2_10_l1 = dados_62213_l1['Informações Complementares 2'].astype(str).str.startswith('10')
    dados_62213_l1 = dados_62213_l1[filtro_info2_10_l1]
    condicao_1031 = dados_62213_l1['Informações Complementares 2'] == "1031"
    dados_62213_l1 = dados_62213_l1[~condicao_1031]
    info5_l1 = dados_62213_l1['Informações Complementares 5'].astype(str)
    condicao_retirar_l1 = (info5_l1.str.len() >= 4) & (info5_l1.str[2] == '9') & (info5_l1.str[3] == '1')
    dados_62213_filtrados_l1 = dados_62213_l1[~condicao_retirar_l1]
    saldo_l1 = dados_62213_filtrados_l1['Saldo Inicial'].sum()
    saldo_l110_formatado = formato_contabil(saldo_l1)

    #D2_00070 L2
    filtro_62213_l2 = arquivo['Conta Contábil'].astype(str).str.startswith('62213')
    dados_62213_l2 = arquivo[filtro_62213_l2]
    filtro_info2_10_l2 = dados_62213_l2['Informações Complementares 2'].astype(str).str.startswith('10')
    dados_62213_l2 = dados_62213_l2[filtro_info2_10_l2]
    condicao_1031 = dados_62213_l2['Informações Complementares 2'] == "1031"
    dados_62213_l2 = dados_62213_l2[~condicao_1031]
    contas_para_remover_l2 = ["622130100", "622130200", "622130500"]
    condicao_remover_contas_l2 = dados_62213_l2['Conta Contábil'].isin(contas_para_remover_l2)
    dados_62213_sem_especificas_l2 = dados_62213_l2[~condicao_remover_contas_l2]
    info5_l2 = dados_62213_sem_especificas_l2['Informações Complementares 5'].astype(str)
    condicao_retirar_l2 = (info5_l2.str.len() >= 4) & (info5_l2.str[2] == '9') & (info5_l2.str[3] == '1')
    dados_62213_filtrados_l2 = dados_62213_sem_especificas_l2[~condicao_retirar_l2]
    saldo_l2 = dados_62213_filtrados_l2['Saldo Inicial'].sum()
    saldo_l210_formatado = formato_contabil(saldo_l2)

    #D2_00071 L1
    filtro_62213_l1 = arquivo['Conta Contábil'].astype(str).str.startswith('62213')
    dados_62213_l1 = arquivo[filtro_62213_l1]
    filtro_info2_10_l1 = dados_62213_l1['Informações Complementares 2'].astype(str).str.startswith('12')
    dados_62213_l1 = dados_62213_l1[filtro_info2_10_l1]
    info5_l1 = dados_62213_l1['Informações Complementares 5'].astype(str)
    condicao_retirar_l1 = (info5_l1.str.len() >= 4) & (info5_l1.str[2] == '9') & (info5_l1.str[3] == '1')
    dados_62213_filtrados_l1 = dados_62213_l1[~condicao_retirar_l1]
    saldo_l1 = dados_62213_filtrados_l1['Saldo Inicial'].sum()
    saldo_l112_formatado = formato_contabil(saldo_l1)
    
    #D2_00071 L2
    filtro_62213_l2 = arquivo['Conta Contábil'].astype(str).str.startswith('62213')
    dados_62213_l2 = arquivo[filtro_62213_l2]
    filtro_info2_10_l2 = dados_62213_l2['Informações Complementares 2'].astype(str).str.startswith('12')
    dados_62213_l2 = dados_62213_l2[filtro_info2_10_l2]
    contas_para_remover_l2 = ["622130100", "622130200", "622130500"]
    condicao_remover_contas_l2 = dados_62213_l2['Conta Contábil'].isin(contas_para_remover_l2)
    dados_62213_sem_especificas_l2 = dados_62213_l2[~condicao_remover_contas_l2]
    info5_l2 = dados_62213_sem_especificas_l2['Informações Complementares 5'].astype(str)
    condicao_retirar_l2 = (info5_l2.str.len() >= 4) & (info5_l2.str[2] == '9') & (info5_l2.str[3] == '1')
    dados_62213_filtrados_l2 = dados_62213_sem_especificas_l2[~condicao_retirar_l2]
    saldo_l2 = dados_62213_filtrados_l2['Saldo Inicial'].sum()
    saldo_l212_formatado = formato_contabil(saldo_l2)

    conta_62213_total_float = formato_float(conta_62213_total_49)
    saldo_l19_formatado_float = formato_float(saldo_l19_formatado)
    saldo_l110_formatado_float = formato_float(saldo_l110_formatado)
    saldo_l112_formatado_float = formato_float(saldo_l112_formatado)

    contas_62213_especificas_float = formato_float(contas_62213_especificas_49)
    saldo_l29_formatado_float = formato_float(saldo_l29_formatado)
    saldo_l210_formatado_float = formato_float(saldo_l210_formatado)
    saldo_l212_formatado_float = formato_float(saldo_l212_formatado)

    conta_72_l1  = conta_62213_total_float-saldo_l19_formatado_float-saldo_l110_formatado_float-saldo_l112_formatado_float
    conta_72_l1 = formato_contabil(conta_72_l1)
    conta_72_l2  = contas_62213_especificas_float-saldo_l29_formatado_float-saldo_l210_formatado_float-saldo_l212_formatado_float
    conta_72_l2 = formato_contabil(conta_72_l2)

    # D2_00073 L1
    contas_desejadas = arquivo['Conta Contábil'].astype(str).str.startswith('62213')
    conta_62213 = arquivo[contas_desejadas]
    info5_l1 = conta_62213['Informações Complementares 5'].astype(str)
    condicao_colocar_l1 = (info5_l1.str.len() >= 4) & (info5_l1.str[2] == '9') & (info5_l1.str[3] == '1')
    conta_62213 = conta_62213[condicao_colocar_l1]
    conta_62213_total_73 = conta_62213['Saldo Inicial'].sum()
    conta_62213_total_73 = formato_contabil(conta_62213_total_73)

    # EXCETO 01 02 E 05 D2_00049 LINHA 2
    contas_desejadas = arquivo[arquivo['Conta Contábil'].astype(str).str.startswith('62213')]
    condicao = contas_desejadas['Conta Contábil'].isin(["622130100", "622130200", "622130500"])
    contas_62213_especificas = contas_desejadas[~condicao]
    info5_l2 = contas_62213_especificas['Informações Complementares 5'].astype(str)
    condicao_colocar_l2 = (info5_l2.str.len() >= 4) & (info5_l2.str[2] == '9') & (info5_l2.str[3] == '1')
    contas_62213_especificas = contas_62213_especificas[condicao_colocar_l2]
    contas_62213_especificas_73 = contas_62213_especificas['Saldo Inicial'].sum()
    contas_62213_especificas_73 = formato_contabil(contas_62213_especificas_73)

    #D2_00074 L1
    contas_desejadas = arquivo['Conta Contábil'].astype(str).str.startswith('6322')
    conta_6322 = arquivo[contas_desejadas]
    conta_6322_total = conta_6322['Saldo Inicial'].sum()
    conta_6322_total = formato_contabil(conta_6322_total)

    #D2_00074 L2
    contas_desejadas = arquivo['Conta Contábil'].astype(str).str.startswith('6314')
    conta_6314 = arquivo[contas_desejadas]
    conta_6314_total = conta_6314['Saldo Inicial'].sum()
    conta_6314_total = formato_contabil(conta_6314_total)

    escrever_valor_abaixo(ws,wb,"D2_00044",conta_6212_6213_total,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,'D2_00045',msc_dezembro_172,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D2_00046",msc_dezembro_111,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D2_00047",saldo_total_formatado,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D2_00048",saldo_total_formatado_48,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D2_00049",conta_62213_total_49,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D2_00049",contas_62213_especificas_49,planilha_conferencia,3)
    escrever_valor_abaixo(ws,wb,"D2_00049",conta_6221304_49,planilha_conferencia,4)
    escrever_valor_abaixo(ws,wb,"D2_00050",conta_6221307,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D2_00050",conta_6221305,planilha_conferencia,3)
    escrever_valor_abaixo(ws,wb,"D2_00058",conta_4522_total,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D2_00058",conta_3522_total,planilha_conferencia,3)
    escrever_valor_abaixo(ws,wb,"D2_00069",saldo_l19_formatado,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D2_00069",saldo_l29_formatado,planilha_conferencia,3)
    escrever_valor_abaixo(ws,wb,"D2_00070",saldo_l110_formatado,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D2_00070",saldo_l210_formatado,planilha_conferencia,3)
    escrever_valor_abaixo(ws,wb,"D2_00071",saldo_l112_formatado,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D2_00071",saldo_l212_formatado,planilha_conferencia,3)
    escrever_valor_abaixo(ws,wb,"D2_00072",conta_72_l1,planilha_conferencia) 
    escrever_valor_abaixo(ws,wb,"D2_00072",conta_72_l2,planilha_conferencia,3)
    escrever_valor_abaixo(ws,wb,"D2_00073",conta_62213_total_73,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D2_00073",contas_62213_especificas_73,planilha_conferencia,3)
    escrever_valor_abaixo(ws,wb,"D2_00074",conta_6322_total,planilha_conferencia)
    escrever_valor_abaixo(ws,wb,"D2_00074",conta_6314_total,planilha_conferencia,3)